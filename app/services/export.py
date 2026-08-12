from __future__ import annotations

import csv
import datetime as dt
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.db.models import Application

CSV_HEADERS = [
    "id",
    "status",
    "category",
    "position",
    "full_name",
    "phone",
    "address",
    "birth_date",
    "work_experience_text",
    "experience_years_range",
    "education_level",
    "education_institution",
    "languages",
    "expected_salary_range",
    "computer_skills",
    "key_skills",
    "source",
    "submitted_at",
]


def _application_row(app_: Application, category_name: str, position_name: str) -> list:
    return [
        app_.id,
        app_.status,
        category_name,
        position_name,
        app_.full_name or "",
        app_.phone or "",
        (app_.address or "").replace("\n", " "),
        app_.birth_date.isoformat() if app_.birth_date else "",
        (app_.work_experience_text or "").replace("\n", " "),
        app_.experience_years_range or "",
        app_.education_level or "",
        app_.education_institution or "",
        ",".join(app_.languages) if app_.languages else "",
        app_.expected_salary_range or "",
        (app_.computer_skills or "").replace("\n", " "),
        (app_.key_skills or "").replace("\n", " "),
        app_.source or "",
        app_.submitted_at.strftime("%Y-%m-%d %H:%M") if app_.submitted_at else "",
    ]


def applications_to_csv(rows: list[tuple[Application, str, str]]) -> str:
    """rows: [(application, category_name_uz, position_name_uz), ...]"""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(CSV_HEADERS)

    for app_, category_name, position_name in rows:
        writer.writerow(_application_row(app_, category_name, position_name))
    return buffer.getvalue()


def applications_to_xlsx(rows: list[tuple[Application, str, str]]) -> bytes:
    """rows: [(application, category_name_uz, position_name_uz), ...]"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    ws.append(CSV_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for app_, category_name, position_name in rows:
        ws.append(_application_row(app_, category_name, position_name))

    for idx, header in enumerate(CSV_HEADERS, start=1):
        max_len = max(
            [len(header)] + [len(str(ws.cell(row=r, column=idx).value or "")) for r in range(2, ws.max_row + 1)]
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_filename(prefix: str = "lazana_applications", ext: str = "csv") -> str:
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{stamp}.{ext}"
